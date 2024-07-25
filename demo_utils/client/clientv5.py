# test client
import time
import typing
import random
from openpyxl import Workbook # type: ignore
from openpyxl import load_workbook
import aiohttp
import asyncio
import os
from datetime import datetime
from pathlib import Path

send_cnt = 0
finished_cnt = 0
loops = 1
requests_batch = 150
client_name = __file__.split("\\")[-1].split(".")[0]
all_requests_sum = loops * requests_batch

is_test = False

def test():
    # TODO test code

    pass

# filename = "response_information_v1(150)"
filename = f"{client_name}#loops{loops}#requests_batch{requests_batch}#{datetime.ctime(datetime.now()).replace(' ', '-').replace(':', '-')}"
dirpath = Path.cwd() / "excel5"

def to_excel(data, filename, dirpath, headers):
    print(headers)
    if not os.path.exists(dirpath):
        os.makedirs(dirpath, exist_ok=True)

    file_path = str(dirpath / f"{filename}.xlsx")

    if os.path.exists(file_path):
        # 如果文件存在，加载工作簿和活动工作表
        workbook = load_workbook(file_path)
        sheet = workbook.active

    else:
        # 如果文件不存在，创建新的工作簿和工作表
        workbook = Workbook()
        sheet = workbook.active
        # 写入表头
        sheet.append(headers)

    for row in data:
        sheet.append(row)

    workbook.save(file_path)


async def fetch(session: aiohttp.ClientSession, url, number):
    global finished_cnt
    global requests_batch
    global send_cnt
    data = {"number": number}
    headers = {"task-type": "C"}

    send_cnt += 1

    print(f"Send count: {send_cnt}/{all_requests_sum}")

    start_time = time.time()

    async with session.post(url, json=data, headers=headers) as response:
        data = await response.json()
        data["total_response_time"] = time.time()  - start_time
        data['trans_delay'] = data['total_response_time'] - data["wait_time_in_worker_node"] - data["process_in_worker_node"]

        finished_cnt += 1
        print(f"process information: {finished_cnt}/{requests_batch}")
        return data


async def main(args):
    host = "192.168.0.100"
    port = 8081
    url = f"http://{host}:{port}"


    tasks = list()
    async with aiohttp.ClientSession(
        connector=aiohttp.TCPConnector(limit=0),
        timeout=aiohttp.ClientTimeout(total=None),
    ) as session:
        # split
        tasks = [fetch(session, url, arg) for arg in args]
        responses = await asyncio.gather(*tasks, return_exceptions=True)

        return responses


async def run():
    # global variable
    global requests_batch
    global filename
    global dirpath
    global finished_cnt

    for _ in range(loops):
        # funciton
        def result_parse(responses: typing.List[typing.Dict[str, typing.Any]]) -> typing.Tuple[int, typing.Dict[str, typing.Any], typing.List]:
            data_table = list()
            response_keys = list()

            for res in responses:
                if type(res) is dict:
                    response_keys = list(res.keys())
                else:
                    raise Exception("Error")

            try:
                if responses:
                    for response in responses:
                        if response.get("success"):
                            data_table.append(
                                [value for key, value in response.items()]
                            )
                        else:
                            data_table.append(["-" for _ in range(len(response_keys))])
                    print("--EXIT--")
                    code = 0
                else:
                    code = 1
                    data_table = None
            except Exception as e:
                print(e)
                code = -1
                data_table = None
            finally:
                return code, data_table, response_keys

        # process
        # args = [random.randint(0, 1000000) for _ in range(requests_sum)]
        args = [500000 for _ in range(requests_batch)]

        print("---start fetch---")

        responses = await main(args)


        for response in responses:
            for k, v in response.items():
                print(f"[{k}]: {v}")

        print("---generate data file---")

        # write into excel file
        code, data_table, col_headers = result_parse(responses)

        if data_table:
            to_excel(data_table, filename, dirpath, col_headers)
        else:
            print("None data_table")

        finished_cnt = 0

        print(f"Cover finished\nExit code: {code}")


if __name__ == "__main__":
    if is_test:
        test()
    else:
        asyncio.run(run())


    pass
