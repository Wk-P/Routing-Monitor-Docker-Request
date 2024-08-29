# test client
import time
import typing
import random
from openpyxl import Workbook  # type: ignore
from openpyxl import load_workbook
import aiohttp
import asyncio
import os
from datetime import datetime
from pathlib import Path
import math

send_cnt = 0
finished_cnt = 0
loops = 1
requests_batch = 100

task_interval = 0.4
batch_interval = 2

client_name = __file__.split("\\")[-1].split(".")[0]
all_requests_sum = loops * requests_batch


random_int_max = 500000
random_int_min = 10


# random request number switch
is_random_request_number = True


# unit code test switch
is_unit_code_test = False


# response console print withou excel
is_test_response_print = False


# single request for test
is_single_request_sum = False

if is_single_request_sum:
    loops = 1
    requests_batch = 50
    all_requests_sum = loops * requests_batch

def test():
    # TODO test code
    print(filename)
    pass


# filename = "response_information_v1(150)"
if is_random_request_number:
    filename = f'''RAND{client_name}-L{loops}-RB{requests_batch}-DT{datetime.ctime(datetime.now()).replace(' ', '').replace(':', '')}'''
else:
    filename = f'''{client_name}-L{loops}-RB{requests_batch}-DT{datetime.ctime(datetime.now()).replace(' ', '').replace(':', '')}'''


if is_single_request_sum:
    filename = f"#test"

dirpath = Path.cwd() / "RS1"


def to_excel(data, filename, dirpath, headers):
    print(headers)
    if not os.path.exists(dirpath):
        os.makedirs(dirpath, exist_ok=True)

    file_path = str(dirpath / f"{filename}.xlsx")

    if os.path.exists(file_path):
        workbook = load_workbook(file_path)
        sheet = workbook.active
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(headers)

    for row in data:
        sheet.append(row)

    workbook.save(file_path)


async def fetch(session: aiohttp.ClientSession, url, number, delay):
    global finished_cnt
    global requests_batch
    global send_cnt
    global loops

    await asyncio.sleep(delay)

    data = {"number": number}
    headers = {"task-type": "C"}

    send_cnt += 1

    print(f"send timestamp: {time.time()}", end='\t')
    print(f"Send count: {send_cnt}/{all_requests_sum}, {round(100 * send_cnt/all_requests_sum, 2)}%")

    start_time = time.time()

    async with session.post(url, json=data, headers=headers) as response:
        data = await response.json()
        data["total_response_time"] = time.time() - start_time
        finished_cnt += 1
        print(f"process timestamp: {time.time()}", end='\t')
        print(f"process information: {finished_cnt}/{all_requests_sum}, {round(100 * finished_cnt/all_requests_sum, 2)}%")
        return data


async def main(args):
    host = "192.168.0.100"
    port = 8081
    url = f"http://{host}:{port}"

    tasks = list()
    responses = list()
    async with aiohttp.ClientSession(
        connector=aiohttp.TCPConnector(limit=0),
        timeout=aiohttp.ClientTimeout(total=None),
    ) as session:
        # split
        for i, arg in enumerate(args):
            delay = i * task_interval
            task = asyncio.create_task(fetch(session, url, arg, delay))
            tasks.append(task)
        response = await asyncio.gather(*tasks, return_exceptions=True)
        responses.append(response)

        return responses

# funciton
def result_parse(responses: typing.List[typing.Dict[str, typing.Any]]) -> typing.Tuple[int, typing.Dict[str, typing.Any], typing.List]:
    data_table = list()
    response_keys = list()

    for res in responses:
        if type(res) is dict:
            response_keys = list(res.keys())
        else:
            raise Exception("Error")

    try:
        if responses:
            for response in responses:
                if response.get("success"):
                    data_table.append(
                        [value for key, value in response.items()]
                    )
                else:
                    data_table.append(
                        ["-" for _ in range(len(response_keys))])
            print("--EXIT--")
            code = 0
        else:
            code = 1
            data_table = None
    except Exception as e:
        print(e)
        code = -1
        data_table = None
    finally:
        return code, data_table, response_keys


async def run():
    # global variable
    global requests_batch
    global filename
    global dirpath
    global finished_cnt
    global is_random_request_number

    for _ in range(loops):
        # process
        if is_random_request_number:
            args = [math.floor(random.uniform(
                random_int_min * 10, random_int_max * 10) / 10) for _ in range(requests_batch)]
        else:
            args = [500000 for _ in range(requests_batch)]

        print("---start fetch---")

        responses = await main(args)

        for response in responses[0]:
            for k, v in response.items():
                if k == "real_wait_time" or k == "predict_wait_time":
                    print(k, v)

        print("---generate data file---")

        # write into excel file
        for response in responses:
            code, data_table, col_headers = result_parse(response)

            if not is_test_response_print:
                if data_table:
                    to_excel(data_table, filename, dirpath, col_headers)
                else:
                    print("None data_table")

                print(f"Cover finished\nExit code: {code}")
            else:
                print(code)


if __name__ == "__main__":
    if is_unit_code_test:
        test()
    else:
        asyncio.run(run())
    pass