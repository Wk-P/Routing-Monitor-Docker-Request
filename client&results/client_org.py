# test client
import re
import time
import typing
import random
from numpy import broadcast_shapes
from openpyxl import Workbook  # type: ignore
from openpyxl import load_workbook
import aiohttp
import asyncio
import os
from datetime import datetime
from pathlib import Path
import math

def read_numbers_from_file():
    with open('args.txt', 'r') as file:
        args_from_file = [int(line.strip()) for line in file]
        return args_from_file
    

def test():
    # TODO test code
    print(client_params.tasks_batches)

    for _il in range(client_params.loops):
        responses_list = list()
        for _ig in range(client_params.group):
            args = list()
            if not client_params.is_read_from_file:
                for _tks in client_params.tasks_batches:
                    # sum of tasks for every group -> _tks
                    if client_params.is_random_request_number:
                        args += [math.floor(random.uniform(
                            client_params.random_int_min * 10, client_params.random_int_max * 10) / 10) for _ in range(_tks)]
                    else:
                        args += [500000 for _ in range(_tks)]
            else:
                args = client_params._args

    print(len(args))
    pass


class ClientParams:
    def __init__(self, *args, **kw) -> None:
        self.send_cnt = 0
        self.finished_cnt = 0
        self.loops = kw.get('loops')
        self.sum_group = kw.get('sum_group')
        self.sum_args_min = kw.get("sum_args_min")
        self.sum_args_max = kw.get("sum_args_max")
        self.groups = [0 for _ in range(self.sum_group)]
        self.task_interval = kw.get('task_interval')
        self.group_interval = kw.get('group_interval')
        self.loops_interval = kw.get('loops_interval')

        for i in range(len(self.groups)):
            # N requests in self.sum_groups
            self.groups[i] = random.randint(self.sum_args_min, self.sum_args_max)

        if self.loops < 2:
            self.loops_interval = 0
        
        if self.group_interval < 2:
            self.group_interval = 0

        if self.task_interval < 2:
            self.task_interval = 0

        self.client_name = __file__.split("\\")[-1].split(".")[0]
        self.all_requests_sum = self.loops * sum(self.groups)

        self.random_int_max = kw.get("random_int_max")
        self.random_int_min = kw.get('random_int_min')

        # random request number switch
        self.is_random_request_number = kw.get("is_random_request_number")

        # unit code test switch
        self.is_unit_code_test = kw.get('is_unit_code_test')

        # response console print withou excel
        self.is_test_response_print = kw.get('is_test_response_print')

        # single request for test
        self.is_single_request_sum = kw.get('is_single_request_sum')

        # request number data from file
        self.is_read_from_file = kw.get('is_read_from_file')


        if self.is_single_request_sum:
            self.loops = 1
            self.groups = [20]
            self.all_requests_sum = self.loops * sum(self.groups)
        
        if self.is_read_from_file:
            self._args = read_numbers_from_file()
            self.groups = len(self._args)

        if self.is_random_request_number:
            self.filename = f'''RAND{self.client_name}-L{self.loops}-G{self.sum_group}-RB{sum(self.groups)}''' + kw.get('filenamekw')
        else:
            self.filename = f'''{self.client_name}-L{self.loops}-G{self.sum_group}-RB{sum(self.groups)}''' + kw.get('filenamekw')

        if self.is_single_request_sum:
            self.filename = f"#test"

        self.dirpath = kw.get('dirpath')

client_params = ClientParams(
    loops = 1,
    sum_group = 200,
    task_interval = 0.3,
    loops_interval = 4,
    group_interval = 5,
    sum_args_min = 0,
    sum_args_max = 500,
    is_random_request_number = True,
    is_unit_code_test = False,
    is_test_response_print = False,
    is_single_request_sum = False,
    is_read_from_file = False,
    dirpath = Path.cwd() / "RSN3",
    random_int_max = 500000,
    random_int_min = 1,
    filenamekw = f'''-DT{datetime.ctime(datetime.now()).replace(' ', '').replace(':', '')}'''
)


def to_excel(data, filename, dirpath, headers):
    print(headers)
    if not os.path.exists(dirpath):
        os.makedirs(dirpath, exist_ok=True)

    file_path = str(dirpath / f"{filename}.xlsx")

    if os.path.exists(file_path):
        workbook = load_workbook(file_path)
        sheet = workbook.active
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(headers)

    for row in data:
        sheet.append(row)

    workbook.save(file_path)


async def fetch(session: aiohttp.ClientSession, url, number):
    global client_params

    data = {"number": number}
    headers = {"task-type": "C"}

    client_params.send_cnt += 1

    print(f"send timestamp: {time.time()}", end='\t')
    print(f"Send count: {client_params.send_cnt}/{client_params.all_requests_sum}, {round(100 * client_params.send_cnt/client_params.all_requests_sum, 2)}%")


    start_time = time.time()

    async with session.post(url, json=data, headers=headers) as response:
        data = await response.json()
        data["total_response_time"] = time.time() - start_time
        client_params.finished_cnt += 1
        print(f"{'start timestamp:':<50}{start_time:<20}")
        print(f"{'process timestamp:':<50}{time.time():<20}", end='\t')
        hint_str = f"{client_params.finished_cnt}/{client_params.all_requests_sum}, {round(100 * client_params.finished_cnt/client_params.all_requests_sum, 2)}%"
        print(f"{'process information:':<50}{hint_str:<20}")
        return data


async def main(args):
    host = "192.168.0.100"
    port = 8081
    url = f"http://{host}:{port}"

    tasks = list()
    responses = list()

    async with aiohttp.ClientSession(
        connector=aiohttp.TCPConnector(limit=0),
        timeout=aiohttp.ClientTimeout(total=None),
    ) as session:
        # split
        for arg in args:
            # delay = i * client_params.task_interval
            task = asyncio.create_task(fetch(session, url, arg))
            tasks.append(task)
            await asyncio.sleep(0.3)

        responses = await asyncio.gather(*tasks)
        return responses

# funciton


def result_parse(responses: typing.List[typing.Dict[str, typing.Any]]) -> typing.Tuple[int, typing.Dict[str, typing.Any], typing.List]:
    data_table = list()
    response_keys = list()

    for res in responses:
        if type(res) is dict:
            response_keys = list(res.keys())
        else:
            raise Exception("Error")

    try:
        if responses:
            for response in responses:
                if response.get("success"):
                    data_table.append(
                        [value for key, value in response.items()]
                    )
                else:
                    data_table.append(
                        ["-" for _ in range(len(response_keys))])
            print("--EXIT--")
            code = 0
        else:
            code = 1
            data_table = None
    except Exception as e:
        print(e)
        code = -1
        data_table = None
    finally:
        return code, data_table, response_keys


async def run():
    # global variable
    global client_params

    for _loop in range(client_params.loops):
        print(f"{'Loop':<40}: {_loop:<20}")
        responses_list = list()
        args = list()
        if not client_params.is_read_from_file:
            for _tks in client_params.groups:
                # sum of tasks for every group -> _tks
                if client_params.is_random_request_number:
                    args = [math.floor(random.uniform(
                        client_params.random_int_min * 10, client_params.random_int_max * 10) / 10) for _ in range(_tks)]
                else:
                    args = [500000 for _ in range(_tks)]
                print("---start fetch---")
                responses = await main(args)
                await asyncio.sleep(client_params.group_interval)
                print("---generate data file---")
                responses_list.append(responses)
        else:
            args = client_params._args

    # write into excel file
    for responses in responses_list:
        code, data_table, col_headers = result_parse(responses)
        if not client_params.is_test_response_print:
            if data_table:
                to_excel(data_table, client_params.filename, client_params.dirpath, col_headers)
            else:
                print("None data_table")
        else:
            print(code)


if __name__ == "__main__":
    if client_params.is_unit_code_test:
        test()
    else:
        PR_start = time.time()
        asyncio.run(run())
        PR_end = time.time()
        print(f"{'PR total time:':<40}{PR_end - PR_start:<20}s")
    pass
