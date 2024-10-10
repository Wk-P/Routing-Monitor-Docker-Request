# test client
from operator import index
import time
import typing
import random
from matplotlib.pyplot import figlegend
from numpy import broadcast_shapes
from openpyxl import Workbook  # type: ignore
from openpyxl import load_workbook
import aiohttp
import asyncio
import os
from datetime import datetime
from pathlib import Path
import logging
import matplotlib.pyplot as plt

import time_graph.generate_graph as figplt

# 로깅 설정: 로그 파일 저장 경로 및 설정
logging.basicConfig(filename = str(Path.cwd() / 'logs' / f'{__file__.split(".")[0]}-output.log'), level=logging.INFO, filemode='w')

# 클라이언트 파라미터 설정 클래스
class ClientParams:
    def __init__(self, *args, **kw) -> None:
        self.send_cnt = 0                                         # 전송된 요청 수
        self.finished_cnt = 0                                     # 처리 완료된 요청 수
        self.requests_sum = kw.get('requests_sum')                # 요청 총합
        self.task_interval = kw.get('task_interval')              # task를 발송하는 간격
        self.random_int_max = kw.get("random_int_max")
        self.random_int_min = kw.get('random_int_min')
        self.group_limit = kw.get('group_limit')
        self.group_interval = kw.get('group_interval')

        # random request number switch
        self.is_random_request_number = kw.get("is_random_request_number")
        # unit code test switch
        self.is_unit_code_test = kw.get('is_unit_code_test')
        # response console print withou excel
        self.is_test_response_print = kw.get('is_test_response_print')
        # single request for test
        self.is_single_request_sum = kw.get('is_single_request_sum')
        # request number data from file
        self.is_read_from_file = kw.get('is_read_from_file')

        # 클라이언트 이름 설정
        self.client_name = __file__.split("\\")[-1].split(".")[0]
        
        if self.is_single_request_sum:
            self.requests_sum = 1

        if self.is_read_from_file:
            self._args = read_numbers_from_file()

        if self.is_random_request_number:
            self.filename = f'''RAND{self.client_name}''' + kw.get('filenamekw')
        else:
            self.filename = f'''{self.client_name}''' + kw.get('filenamekw')

        if self.is_single_request_sum:
            self.filename = f"#test"

        self.dirpath = kw.get('dirpath')

# 파일에서 숫자 불러오기
def read_numbers_from_file():
    with open('args.txt', 'r') as file:
        args_from_file = [int(line.strip()) for line in file]
        return args_from_file
    

def test():
    # TODO test code
    pass


REQUESTS_SUM = 100
LOOPS = 20
client_params = ClientParams(
    task_interval = 0.03,
    requests_sum = REQUESTS_SUM,
    group_limit = 1,
    group_interval = 0.5,
    is_random_request_number = True,
    is_unit_code_test = False,
    is_test_response_print = False,
    is_single_request_sum = False,
    is_read_from_file = False,
    dirpath = Path.cwd() / f"{REQUESTS_SUM}_train_data",
    random_int_max = 300000,
    random_int_min = 1,
    filenamekw = f'''-DT{datetime.ctime(datetime.now()).replace(' ', '').replace(':', '')}'''
)

# 데이터를 엑셀 파일로 저장하는 함수
def to_excel(data, filename, dirpath, headers):
    if not os.path.exists(dirpath):
        os.makedirs(dirpath, exist_ok=True)

    file_path = str(dirpath / f"{filename}.xlsx")

    # 파일이 이미 존재하는 경우 불러오기, 그렇지 않으면 새로 생성
    if os.path.exists(file_path):
        workbook = load_workbook(file_path)
        sheet = workbook.active
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(headers)     # 헤더 추가

    # 데이터 추가
    for row in data:
        sheet.append(row)

    workbook.save(file_path)

# 요청을 보내는 비동기 함수
async def fetch(session: aiohttp.ClientSession, url, number, task_index):
    global client_params

    data = {"number": number}         # 요청 데이터
    headers = {"task-type": "C"}      # 요청 헤더

    client_params.send_cnt += 1       # 전송 카운트 증가

    logging.info(f"send timestamp: {time.time()} \t")     # 전송 시간 로깅
    logging.info(f"Send count: {client_params.send_cnt}/{client_params.requests_sum}, {round(100 * client_params.send_cnt/client_params.requests_sum, 2)}%\n")

    start_time = time.time()          # 시작 시간을 기록
    try:
        response_data = dict()
        # 비동기로 요청 보내기
        async with session.post(url, json=data, headers=headers) as response:
            print(f'Reponse status: {response.status}')
            response_data = await response.json()                             # 응답을 JSON 형태로 받기
            response_data["total_response_time"] = time.time() - start_time   # 총 응답 시간을 계산
            response_data['task_index'] = task_index                          # 작업 인덱스 기록
            client_params.finished_cnt += 1                                   # 완료 카운트 증가
            logging.info(f"{'start timestamp:':<50}{start_time:<20}\n")
            logging.info(f"{'process timestamp:':<50}{time.time():<20}\t")
            hint_str = f"{client_params.finished_cnt}/{client_params.requests_sum}, {round(100 * client_params.finished_cnt/client_params.requests_sum, 2)}%"
            logging.info(f"{'process information:':<50}{hint_str:<20}\n")
            return response_data
    except Exception as e:
        print(e)
        print('*' * 20, datetime.ctime(datetime.now()), '*' * 20)


async def main(args):
    host = "192.168.0.100"              # 서버 호스트
    port = 8100                         # 서버 포트
    url = f"http://{host}:{port}"

    tasks = list()                      # 비동기 작업 리스트
    responses = list()                  # 응답 리스트

    # 비동기 클라이언트 세션 설정
    async with aiohttp.ClientSession(
        connector=aiohttp.TCPConnector(limit=0),
        timeout=aiohttp.ClientTimeout(total=None),
    ) as session:
        # split
        _index = 0
        # 요청마다 비동기 task 생성
        for arg in args:
            # if _index == client_params.group_limit:
            #     _index = 0
            #     await asyncio.sleep(client_params.group_interval)

            task = asyncio.create_task(fetch(session, url, arg, _index))
            tasks.append(task)
            await asyncio.sleep(client_params.task_interval)    # task 간격만큼 대기
            print(_index)
            _index += 1

        responses = await asyncio.gather(*tasks, return_exceptions=True)   # 모든 task를 모아서 실행
        return responses


# 결과를 처리하고 엑셀 파일로 저장할 데이터를 파싱하는 함수
def result_parse(responses: typing.List[typing.Dict[str, typing.Any]]) -> typing.Tuple[int, typing.Dict[str, typing.Any], typing.List]:
    data_table = list()        # 데이터 테이블
    response_keys = list()     # 응답 키

    # 응답 파싱
    for res in responses:
        if type(res) is dict:
            response_keys = list(res.keys())
        else:
            raise Exception("Error")

    try:
        if responses:
            for response in responses:
                if response.get("success"):
                    data_table.append(
                        [value for key, value in response.items()]
                    )
                else:
                    data_table.append(
                        ["-" for _ in range(len(response_keys))])   # 실패 시 빈 데이터 추가
            print("--EXIT--")
            code = 0
        else:
            code = 1
            data_table = None
    except Exception as e:
        print(e)
        code = -1
        data_table = None
    finally:
        return code, data_table, response_keys

# 보상 값을 위한 전역 변수
rewards = 0
# 실행 함수
async def run():
    global pic_index, rewards
    ORG_start = time.time()


    # global variable
    global client_params
    args = list()                # 요청 인자를 담을 리스트
    if not client_params.is_read_from_file:
        # sum of tasks for every group -> _tks    각 그룹에 대한 task 생성
        if client_params.is_random_request_number:
            args = [random.randint(client_params.random_int_min, client_params.random_int_max) for _ in range(client_params.requests_sum)]
        else:
            args = [500000 for _ in range(client_params.requests_sum)]
        
    else:
        args = client_params._args
    
    print("---start fetch---")
    responses = await main(args)  # 비동기 요청 실행
    print("---generate data file---")
    
    # 응답에서 각 데이터 추출
    real_total = []
    pred_total = []
    processed_time = []
    before_forward_time = []
    real_task_wait_time = []
    pred_task_wait_time = []
    before_forward_timestamps = []
    start_process_timestamps = []
    all_rewards_list = []



    end_line = f"\n{'-' * 40}\n"
    for response in responses:
        for k, v in response.items():
            print(k ,v)
            if k == 'total_response_time':
                print(k ,v)
                real_total.append(v)
            elif k == 'total_response_time_prediction':
                print(k ,v)
                pred_total.append(v)
            elif k == 'before_forward_time':
                print(k ,v)
                before_forward_time.append(v)
            elif k == 'real_task_wait_time':
                print(k ,v)
                real_task_wait_time.append(v)
            elif k == 'pred_task_wait_time':
                print(k ,v)
                pred_task_wait_time.append(v)    
            elif k == 'before_forward_timestamp':
                print(k ,v)
                before_forward_timestamps.append(v % 1000)
            elif k == 'start_process_time':
                print(k ,v)
                start_process_timestamps.append(v % 1000)
            elif k == 'processed_time':
                print(k ,v)
                processed_time.append(v)
            elif k == 'rewards':
                print(k ,v)
                all_rewards_list.append(v)
                rewards = sum(all_rewards_list)
            else:
                pass
    print(end_line)
    
    # 결과 출력
    print(real_total)
    print(pred_total)
    print(processed_time)
    print(before_forward_time)
    print(real_task_wait_time)
    print(pred_task_wait_time)
    print(before_forward_timestamps)
    print(start_process_timestamps)

    ORG_end = time.time()    # 종료 시간을 기록
    print(f"{'OGR total time:':<40}{ORG_end - ORG_start:<20}s")

    
    
    
    # figplt.main([[figplt.Data(real_total, 'real total'), figplt.Data(pred_total, 'pred total'), figplt.Data(processed_time, 'process')], [figplt.Data(real_task_wait_time, 'real wait time'), figplt.Data(pred_task_wait_time, 'pred wait time') ], 
    #             [ figplt.Data(start_process_timestamps, 'SPT-ST'), figplt.Data(before_forward_timestamps, 'BFT-ST')]], 
    #             ['real total - pred total - process', 'real - pred (task wait time)', 'start - before'], fig_name=pic_index, fig_dir_path=pic_dir_path,
    #             direction='column')
    

    

    pic_index += 1   # 그래프 이미지 인덱스 증가


    # write into excel file
    code, data_table, col_headers = result_parse(responses)

    if not client_params.is_test_response_print:
        if data_table:
            to_excel(data_table, client_params.filename, client_params.dirpath, col_headers)
        else:
            print("None data_table")
    else:
        print(code)



# pic_index = 1
# pic_dir_path = Path.cwd() / 'figs' / 'fig2'
pic_index = 0
pic_dir_path = None
if __name__ == "__main__":
    if client_params.is_unit_code_test:
        test()
    else:
        all_rewards = []
        for i in range(LOOPS):
            asyncio.run(run())
            all_rewards.append(rewards)
           
        # 보상 값을 그래프로 출력
        plt.plot(all_rewards)
        plt.show()
    pass